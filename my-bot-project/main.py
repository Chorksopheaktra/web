import gspread
from oauth2client.service_account import ServiceAccountCredentials
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
from openpyxl import Workbook, load_workbook
import os

# 🔧 ព័ត៌មានសំខាន់ៗ
SHEET_ID = '19gr_7c9ZcMiPc1AFeXct9PRt_ihon9JfZv4FBLYebVU'
JSON_KEY_FILE = 'service_account.json'
BOT_TOKEN = '7611461271:AAFam8n-96EywhICHBrxbgB-Y0Cw1exIevA'
EXCEL_FILE = 'user_inputs.xlsx'
SHEET_NAME = 'Inputs'

# 🗂️ តភ្ជាប់ទៅ Google Sheet
def connect_sheet():
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    creds = ServiceAccountCredentials.from_json_keyfile_name(JSON_KEY_FILE, scope)
    client = gspread.authorize(creds)
    sheet = client.open_by_key(SHEET_ID).sheet1
    return sheet

# 💾 រក្សាទុកទៅ Excel
def save_input_to_excel(user_input):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME
        ws.append(["User Input"])
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb[SHEET_NAME]
    ws.append([user_input])
    wb.save(EXCEL_FILE)

# 📝 រក្សាទុក input ទៅ Google Sheet (Column A)
def save_to_google_sheet(user_input):
    sheet = connect_sheet()
    sheet.append_row([user_input])  # បញ្ចូលទៅជួរដេកថ្មី (Column A)

# 📥 ដំណើរការបញ្ចូលរបស់អ្នកប្រើ
async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_input = update.message.text.strip().replace(" ", "").lower()

    # ✅ Save to Excel & Google Sheet
    save_input_to_excel(user_input)
    save_to_google_sheet(user_input)

    # 🔍 ស្វែងរកក្នុង Google Sheet
    sheet = connect_sheet()
    records = sheet.get_all_records()

    found = None
    print("📥 User input:", user_input)

    for record in records:
        student_id = str(record.get("អត្តលេខ", "")).strip().replace(" ", "").lower()
        name = str(record.get("ឈ្មោះ", "")).strip().replace(" ", "").lower()
        index = str(record.get("ល.រ", "")).strip().replace(" ", "").lower()

        print("📄 Comparing to:", student_id, "|", name, "|", index)

        if user_input in [student_id, name, index]:
            found = record
            break

    if found:
        reply = "🔍 ទិន្នន័យសិស្ស:\n\n" + "\n".join([f"{k}: {v}" for k, v in found.items()])
        await update.message.reply_text(reply)
    else:
        await update.message.reply_text("❌ មិនមានទិន្នន័យដែលត្រូវនឹងការស្វែងរកនោះទេ។")

# ▶ /start Command
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("📥 សូមបញ្ចូល *អត្តលេខ*, *ឈ្មោះ* ឬ *ល.រ* របស់សិស្ស៖")

# ▶ Main Function
def main():
    app = ApplicationBuilder().token(BOT_TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    print("🤖 Bot កំពុងដំណើរការ...")
    app.run_polling()

if __name__ == '__main__':
    main()
